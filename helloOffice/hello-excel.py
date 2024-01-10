# _*_ coding : utf-8 _*_
# @Time : 2023-10-01 16:30
# @Author : xcsg
# @File :hello-excel
import openpyxl
import re


def modify(file_name: str):
    # 打开文件
    workbook = openpyxl.load_workbook(filename=f"{file_name}.xlsx")
    # 选择工作表
    sheet = workbook.active
    # # 检查是否存在新的一列
    # for column in sheet.columns:
    #     print(column)
    # for row in sheet.iter_rows(min_row=1):
    #     print(row)
    # if 'new_column' not in sheet.columns:
    #     # 在工作表中新增一列
    #     sheet.insert_cols(2)
    for row in sheet.iter_rows(min_row=1):  # 从第1行开始遍历
        cell_value = row[0].value
        # print("\t", row[0].value, "\t", row[1].value)
        # 使用正则表达式提取数字，匹配任何数字（整数或小数）
        matches = re.findall(r'\d+\.?\d*', str(cell_value))
        if matches:
            # 计算乘积
            product = 1
            for match in matches:
                if '.' in match:  # 处理小数
                    product *= float(match)
                else:  # 处理整数
                    product *= int(match)
            # 计算乘积并插入到第三列
            if row[1].value is not None:
                col2 = float(row[1].value)
            else:
                # 处理None值的情况，例如赋予默认值
                col2 = 1.0
            product *= col2
            # 打印表格
            print("\t", row[0].value, "\t", row[1].value, "\t", product)
            # 将结果写入第三列
            row[2].value = round(product / 1000000, 5)

    # 保存修改后的文件
    workbook.save(f"{file_name}_modified.xlsx")


if __name__ == '__main__':
    modify("副本康康救救")
